import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)
CORS(app)

# ==========================================
# 1. LÓGICA DE EXCEL
# ==========================================
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


def extraer_dia(valor_fecha):
    if valor_fecha is None: return ""
    if isinstance(valor_fecha, datetime): return str(valor_fecha.day)
    elif isinstance(valor_fecha, str) and "/" in valor_fecha: return valor_fecha.split("/")[0]
    return str(valor_fecha).strip()


def extraer_mes(valor_fecha):
    if valor_fecha is None: return ""
    if isinstance(valor_fecha, datetime): return MESES_ES.get(valor_fecha.month, "")
    elif isinstance(valor_fecha, str) and "/" in valor_fecha:
        partes = valor_fecha.split("/")
        if len(partes) >= 2:
            try: return MESES_ES.get(int(partes[1]), "")
            except ValueError: return ""
    return ""


def tiene_color(celda):
    if celda.fill and celda.fill.patternType == 'solid':
        color = celda.fill.start_color.index
        if color not in ['00000000', 'FFFFFFFF', 0, '00FFFFFF']: return True
    return False


@app.route('/')
def index():
    html_path = os.path.join(os.path.dirname(__file__), 'index.html')
    return send_file(html_path)


@app.route('/process', methods=['POST'])
def process():
    if 'file' not in request.files: return jsonify({'error': 'No se envió ningún archivo'}), 400
    file = request.files['file']
    if not file.filename.endswith('.xlsx'): return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    try:
        wb_origen = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        lista_principal = []
        lista_ceses = []

        for nombre_hoja in ['Data', 'Ingresos']:
            if nombre_hoja in wb_origen.sheetnames:
                hoja = wb_origen[nombre_hoja]
                for fila in range(2, hoja.max_row + 1):
                    celda_c = hoja.cell(row=fila, column=3)
                    celda_f = hoja.cell(row=fila, column=6)
                    celda_j = hoja.cell(row=fila, column=10)
                    if tiene_color(celda_c) or tiene_color(celda_f) or tiene_color(celda_j): continue
                    nombre = celda_c.value
                    if nombre:
                        lista_principal.append((
                            str(nombre).strip(), str(extraer_dia(celda_f.value)),
                            extraer_mes(celda_f.value), str(celda_j.value or '').strip()
                        ))

        if 'Ceses' in wb_origen.sheetnames:
            hoja_ceses = wb_origen['Ceses']
            for fila in range(2, hoja_ceses.max_row + 1):
                nombre = hoja_ceses.cell(row=fila, column=3).value
                if nombre:
                    fecha_c = hoja_ceses.cell(row=fila, column=6).value
                    lista_ceses.append((
                        str(nombre).strip(), str(extraer_dia(fecha_c)),
                        extraer_mes(fecha_c), str(hoja_ceses.cell(row=fila, column=10).value or '').strip()
                    ))

        lista_final = [r for r in lista_principal if r not in lista_ceses]

        wb_nuevo = openpyxl.Workbook()
        ws = wb_nuevo.active
        assert ws is not None
        ws.title = "Datos Consolidados"
        ws.views.sheetView[0].showGridLines = True

        encabezados = ["Nombres y Apellidos", "Día", "Mes", "Columna J"]
        ws.append(encabezados)

        fuente_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        fill_header   = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        fill_cebra    = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        fuente_datos  = Font(name="Arial", size=10)
        alin_centro   = Alignment(horizontal="center", vertical="center")
        alin_izq      = Alignment(horizontal="left",   vertical="center")
        borde = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9')
        )

        ws.row_dimensions[1].height = 26
        for col_num in range(1, len(encabezados) + 1):
            c = ws.cell(row=1, column=col_num)
            c.font = fuente_header
            c.fill = fill_header
            c.alignment = alin_centro

        for i, registro in enumerate(lista_final, start=2):
            ws.row_dimensions[i].height = 20
            for j, valor in enumerate(registro, start=1):
                c = ws.cell(row=i, column=j, value=valor)
                c.font = fuente_datos
                c.border = borde
                c.alignment = alin_centro if j in (2, 3) else alin_izq
                if i % 2 == 0: c.fill = fill_cebra

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            assert col[0].column is not None
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, 15)

        output = io.BytesIO()
        wb_nuevo.save(output)
        output.seek(0)
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='Resultado_Patron.xlsx')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==========================================
# 2. MOTOR DE IMÁGENES
# ==========================================
def dibujar_texto_rotado(fondo, texto, posicion, angulo, fuente, color=(255, 255, 255, 255)):
    # Canvas 4x para supersampling: dibujamos grande y reducimos con LANCZOS
    escala = 4
    canvas_w, canvas_h = 1200, 400
    img_txt = Image.new('RGBA', (canvas_w * escala, canvas_h * escala), (255, 255, 255, 0))
    d = ImageDraw.Draw(img_txt)
    fuente_4x = fuente.font_variant(size=fuente.size * escala)
    d.text((40, 40), texto, font=fuente_4x, fill=color)
    rotado_4x = img_txt.rotate(angulo, expand=1, resample=Image.Resampling.BICUBIC)
    rotado = rotado_4x.resize(
        (rotado_4x.width // escala, rotado_4x.height // escala),
        Image.Resampling.LANCZOS
    )
    fondo.paste(rotado, posicion, rotado)


@app.route('/generar-tarjeta', methods=['GET'])
def generar_tarjeta():
    # ── Datos del cumpleañero ──────────────────────────────────────
    dia    = request.args.get('dia',    '').upper()
    mes    = request.args.get('mes',    '').upper()
    nombre = request.args.get('nombre', '').title()
    n1     = request.args.get('N1',     '').title()
    area   = request.args.get('area',   '').title()

    # ── Helper: lee parámetro int de la URL, con default ──────────
    def ip(key, default):
        return int(request.args.get(key, default))

    try:
        ruta_fondo = os.path.join(os.path.dirname(__file__), 'static', 'pCumple.jpg')
        fondo = Image.open(ruta_fondo).convert("RGBA")

        # Usar resolución original (1280x720) para máxima calidad
        _, img_h = fondo.size
        draw = ImageDraw.Draw(fondo)

        cy = img_h // 2  # centro vertical = 360px

        # ── Fuentes ────────────────────────────────────────────────
        ruta_fuente      = os.path.join(os.path.dirname(__file__), 'arial.ttf')
        ruta_fuente_bold = os.path.join(os.path.dirname(__file__), 'arialbd.ttf')
        try:
            f_mes    = ImageFont.truetype(ruta_fuente_bold, ip('fmes',    15))
            f_dia    = ImageFont.truetype(ruta_fuente_bold, ip('fdia',    23))
            f_nombre = ImageFont.truetype(ruta_fuente,      ip('fnombre', 27))
            f_area   = ImageFont.truetype(ruta_fuente,      ip('farea',   27))
            f_n1     = ImageFont.truetype(ruta_fuente,      ip('fn1',     20))
        except IOError:
            f_mes = f_dia = f_nombre = f_area = f_n1 = ImageFont.load_default()

        # Ángulo de rotación — CSS rotate(-14deg) = Pillow rotate(14)
        angulo = ip('angulo', 14)

        # ── Box de centrado para nombre, área y N1 ───────────────────
        # &box_x1= &box_x2= definen el área horizontal donde se centra el texto
        # ── Box de centrado (x1=623, x2=1091) medido sobre la imagen real ──
        box_x1 = ip('box_x1', 623)
        box_x2 = ip('box_x2', 1091)

        def cx(texto, fuente):
            ancho = draw.textlength(texto, font=fuente)
            return box_x1 + (box_x2 - box_x1 - int(ancho)) // 2

        # ── Posiciones: Y fijo por URL, X centrado dentro del box ─────────
        dibujar_texto_rotado(fondo, mes,
            posicion=(ip('xmes', 815), cy + ip('ymes', -525)),
            angulo=angulo, fuente=f_mes, color=(255, 255, 255, 255))

        dibujar_texto_rotado(fondo, dia,
            posicion=(ip('xdia', 830), cy + ip('ydia', -510)),
            angulo=angulo, fuente=f_dia, color=(0, 0, 0, 255))

        draw.text((cx(nombre, f_nombre), cy + ip('ynombre', -58)),
            nombre, font=f_nombre, fill=(255, 255, 255, 255))

        draw.text((cx(area, f_area), cy + ip('yarea', -21)),
            area, font=f_area, fill=(255, 255, 255, 255))

        draw.text((ip('xn1', 780), cy + ip('yn1', 21)),
            n1, font=f_n1, fill=(255, 255, 255, 255))

        output = io.BytesIO()
        fondo.save(output, format="PNG")
        output.seek(0)
        response = send_file(output, mimetype='image/png')
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        return response

    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)
